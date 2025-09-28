from pathlib import Path
import pandas as pd
from report_utils import (
    load_pars, load_book,
    parse_raw, prepare_grouped, prepare_logistics
)
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


def format_excel(out_file: Path, min_width: int = 10, max_width: int = 80, debug: bool = True):
    """
    Надёжное форматирование итогового Excel:
      - создаёт Excel Table (если возможно) -> таблицы показывают фильтры в Excel,
      - fallback: ставит ws.auto_filter.ref,
      - делает перенос текста в заголовках, жирный шрифт,
      - автоширина колонок (с ограничением min/max),
      - фиксирует первую строку (A2).
    debug=True выводит полезную диагностику в консоль.
    """
    wb = load_workbook(out_file)
    for idx, ws in enumerate(wb.worksheets, start=1):
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0

        if max_row < 1 or max_col < 1:
            if debug:
                print(f"[format_excel] sheet '{ws.title}' пустой — пропускаем")
            continue

        last_col = get_column_letter(max_col)
        data_ref = f"A1:{last_col}{max_row}"

        # --- Шапка: перенос текста и жирный шрифт ---
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=1, column=col_idx)
            # сохраняем прежние align, ставим wrap_text=True
            prev_align = cell.alignment if cell.alignment is not None else Alignment()
            cell.alignment = Alignment(
                wrap_text=True,
                horizontal=prev_align.horizontal or "center",
                vertical=prev_align.vertical or "center"
            )
            cell.font = Font(bold=True)

        # --- Попытка добавить Excel Table (это включает фильтры в UI) ---
        table_added = False
        table_name = f"Tbl{idx}"
        # displayName должен быть уникален, начинаться с буквы и без пробелов — используем Tbl1, Tbl2...
        try:
            table = Table(displayName=table_name, ref=data_ref)
            style = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            table.tableStyleInfo = style
            # удалим таблицу с таким именем, если случайно уже есть (редко)
            existing = [t.displayName for t in ws._tables]
            if table_name in existing:
                # если уже есть, генерируем имя с суффиксом
                table_name = f"Tbl{idx}_{len(existing)+1}"
                table.displayName = table_name
            ws.add_table(table)
            table_added = True
        except Exception as e:
            # fallback: поставить автофильтр рефом диапазона
            try:
                ws.auto_filter.ref = data_ref
            except Exception:
                # ничего не делаем, просто логируем
                pass
            if debug:
                print(f"[format_excel] не удалось создать Table на листе '{ws.title}': {e}. Поставили auto_filter.ref вместо этого.")

        # --- Фиксируем первую строку (шапку) ---
        try:
            ws.freeze_panes = "A2"
        except Exception:
            # игнорируем, если не получится
            pass

        # --- Автоширина колонок (по всем строкам, первая строка учитывается) ---
        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for row_idx in range(1, max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                val = cell.value
                if val is None:
                    continue
                # для чисел/дат аккуратно
                if isinstance(val, float):
                    s = f"{val:.2f}"
                else:
                    s = str(val)
                # убираем многострочные переносы влияющие на ширину
                s = s.replace("\r", " ").replace("\n", " ")
                if len(s) > max_len:
                    max_len = len(s)
            width = max(min_width, min(max_len + 2, max_width))
            try:
                ws.column_dimensions[col_letter].width = width
            except Exception:
                # возможны ошибки для некоторых специальных столбцов — игнорируем
                pass

        if debug:
            print(f"[format_excel] sheet '{ws.title}': rows={max_row}, cols={max_col}, ref='{data_ref}', table_added={table_added}")

    wb.save(out_file)
    if debug:
        print(f"[format_excel] сохранён {out_file}")


def main():
    root = Path(__file__).parent
    final_dir = root / "final"
    final_dir.mkdir(exist_ok=True)

    pars_cols = load_pars(root)
    book = load_book(root)

    # Оригинальный отчёт
    pre_files = sorted((root / "pre").glob("*.xlsx"))
    if not pre_files:
        raise FileNotFoundError("Нет файлов в папке pre/")
    src = pre_files[0]
    original = pd.read_excel(src)

    # Short
    short = parse_raw(src, pars_cols)

    #checkpoint
    print("Колонки в short:", list(short.columns))

    # Grouped
    grouped = prepare_grouped(short, book)

    # Логистика
    logistics = prepare_logistics(original, book)

    # Запись в Excel
    out_file = final_dir / "final_report.xlsx"
    with pd.ExcelWriter(out_file, engine="openpyxl") as writer:
        original.to_excel(writer, sheet_name="Оригинальный отчет", index=False)
        short.to_excel(writer, sheet_name="Short", index=False)
        grouped.to_excel(writer, sheet_name="Grouped", index=False)
        logistics.to_excel(writer, sheet_name="Логистика", index=False)
    format_excel(out_file)

    print("✅ Отчёт сохранён в", out_file)


if __name__ == "__main__":
    main()
